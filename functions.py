import pandas as pd
from openpyxl import Workbook
from tkinter import filedialog, messagebox
from Cliente_model import Cliente
isencao = []
isentos = []


def formatar(valor):
    return "{:.0f}".format(valor)


def isencao_cartao(arquivo, diretorio):
    planilha = pd.read_excel(arquivo)
    planilha = planilha.dropna()
    planilha["cliente_cpf"] = planilha["cliente_cpf"].apply(formatar)
    planilha["numero_cartao"] = planilha["numero_cartao"].apply(formatar)

    lista = pd.DataFrame(planilha)
    # print(lista)

    for linha, obj in lista.iterrows():
        cliente = Cliente(cliente_cpf=obj['cliente_cpf'], numero_cartao=obj['numero_cartao'],
                          bandeira_cartao=obj['bandeira_cartao'], gasto_mensal=obj['gasto_mensal'])
        isencao.append(cliente)

    for cliente in isencao:
        if cliente.gasto_mensal > 4000:
            isentos.append(cliente.show_all_information())
        else:
            continue

    workbook = Workbook()
    worksheet = workbook.active

    for i, item in enumerate(isentos, start=1):
        item_split = item.split(',')
        for j, valor in enumerate(item_split, start=1):
            chave, valor = valor.split(':')
            worksheet.cell(row=i, column=j).value = valor

    nomes_colunas = ['CPF', 'Número do Cartão', 'Bandeira do Cartão', 'Gasto Mensal']
    for col, nome in enumerate(nomes_colunas, start=1):
        worksheet.cell(row=1, column=col).value = nome

    save_file = diretorio + "/doc.xlsx"
    workbook.save(save_file)
    messagebox.showinfo("Concluído", "Arquivo salvo com sucesso!")
    return None


def start():
    while True:
        messagebox.askyesno(title="Isenção", message="Selecione o arquivo excel para realizar a isenção.")
        arquivo = filedialog.askopenfilename()

        if not arquivo:
            break

        messagebox.askyesno(title="Isenção", message="Selecione o diretório aonde deseja salvar.")
        diretorio = filedialog.askdirectory()

        if not diretorio:
            break

        isencao_cartao(arquivo, diretorio)
        break

    return None
